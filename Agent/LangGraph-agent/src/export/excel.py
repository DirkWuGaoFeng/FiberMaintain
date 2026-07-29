"""
Excel export engine using openpyxl.
"""

from __future__ import annotations

import os
import uuid


def generate_excel(title: str, data: list[dict], sheet_name: str = "Sheet1", output_dir: str = "") -> str:
    """Generate an Excel report. Returns the output file path."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    output_dir = output_dir or "/tmp/fiber_reports"
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, f"{uuid.uuid4().hex}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Title row
    ws.append([title])
    ws["A1"].font = Font(bold=True, size=14)

    if data:
        ws.append([])  # Blank row
        headers = list(data[0].keys())
        ws.append(headers)
        # Bold headers
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=3, column=col_idx).font = Font(bold=True)

        for row in data:
            ws.append([row.get(h, "") for h in headers])

    wb.save(output_path)
    return output_path
