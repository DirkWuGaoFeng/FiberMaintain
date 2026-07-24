"""导出引擎：PDF（reportlab）/ Excel（openpyxl）/ CSV。
文件：/tmp/fiber_reports/{trace_id}_{timestamp}.{ext}，24h 自动清理。
"""
from __future__ import annotations
import asyncio, csv, time, uuid
from datetime import datetime, timezone
from pathlib import Path

from src.settings import settings
from src.mcp.fiber_backend import trace_id_var
from src.monitoring.metrics import REPORT_FILES

REPORT_DIR = Path(settings.app["report_dir"])
RETENTION_S = settings.app["report_retention_hours"] * 3600

# 注册表：file_id → 路径
_FILE_REGISTRY: dict[str, Path] = {}


def _register(path: Path) -> dict:
    file_id = uuid.uuid4().hex[:12]
    _FILE_REGISTRY[file_id] = path
    REPORT_FILES.set(len(_FILE_REGISTRY))
    expires = datetime.now(timezone.utc).astimezone().timestamp() + RETENTION_S
    return {
        "file_id": file_id,
        "filename": path.name,
        "download_url": f"/api/v1/reports/download/{file_id}",
        "expires_at": datetime.fromtimestamp(
            expires, tz=timezone.utc).astimezone().isoformat(),
        "note": "链接 24 小时内有效",
    }


def resolve_file(file_id: str) -> Path | None:
    p = _FILE_REGISTRY.get(file_id)
    if p and p.exists():
        return p
    # 注册表丢失时按目录扫描兜底
    for f in REPORT_DIR.iterdir():
        if f.stem.endswith(file_id):
            return f
    return None


def cleanup_expired() -> int:
    now = time.time()
    removed = 0
    for f in REPORT_DIR.iterdir():
        if now - f.stat().st_mtime > RETENTION_S:
            f.unlink(missing_ok=True)
            removed += 1
    _FILE_REGISTRY.clear()
    return removed


# ───────────────────────── PDF ─────────────────────────
def _build_pdf(title: str, sections: list[dict],
               chart: dict | None, path: Path) -> None:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    from reportlab.platypus import (SimpleDocTemplate, Paragraph,
                                    Spacer, Table, TableStyle, Image)
    pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))

    styles = getSampleStyleSheet()
    zh_title = ParagraphStyle("zhTitle", parent=styles["Title"],
                              fontName="STSong-Light", fontSize=20)
    zh_h = ParagraphStyle("zhH", parent=styles["Heading2"],
                          fontName="STSong-Light", fontSize=13,
                          spaceBefore=10)
    zh_body = ParagraphStyle("zhBody", parent=styles["Normal"],
                             fontName="STSong-Light", fontSize=9.5,
                             leading=14)

    doc = SimpleDocTemplate(str(path), pagesize=A4,
                            topMargin=18 * mm, bottomMargin=18 * mm)
    story = [
        Paragraph(title, zh_title),
        Paragraph(f"生成时间：{datetime.now().astimezone().isoformat()}"
                  f"　|　光纤维护服务系统 Agent", zh_body),
        Spacer(1, 8 * mm),
    ]

    for sec in sections:
        story.append(Paragraph(sec.get("heading", ""), zh_h))
        if "body" in sec and sec["body"]:
            story.append(Paragraph(str(sec["body"]).replace("\n", "<br/>"),
                                   zh_body))
        tbl = sec.get("table")
        if tbl:
            data = [tbl["headers"]] + tbl["rows"]
            t = Table(data, hAlign="LEFT")
            t.setStyle(TableStyle([
                ("FONTNAME", (0, 0), (-1, -1), "STSong-Light"),
                ("FONTSIZE", (0, 0), (-1, -1), 8.5),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1d3557")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#adb5bd")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1),
                 [colors.white, colors.HexColor("#f1f3f5")]),
            ]))
            story.append(t)
        story.append(Spacer(1, 3 * mm))

    if chart:
        img = _render_chart_image(chart, path.with_suffix(".png"))
        if img:
            story.append(Paragraph(chart.get("title", "趋势图"), zh_h))
            story.append(Image(str(img), width=170 * mm, height=80 * mm))

    doc.build(story)


def _render_chart_image(chart: dict, path: Path) -> Path | None:
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        plt.rcParams["font.sans-serif"] = ["Noto Sans CJK SC",
                                           "WenQuanYi Micro Hei", "DejaVu Sans"]
        fig, ax = plt.subplots(figsize=(10, 4.5), dpi=120)
        x = chart.get("x", [])
        for s in chart.get("series", []):
            ax.plot(x, s["data"], marker="o", markersize=3,
                    label=s.get("name", ""))
        ax.set_title(chart.get("title", ""))
        ax.legend(); ax.grid(alpha=0.3)
        fig.autofmt_xdate() if x else None
        fig.tight_layout(); fig.savefig(path)
        plt.close(fig)
        return path
    except Exception:
        return None


# ───────────────────────── Excel ─────────────────────────
def _build_excel(title: str, sections: list[dict],
                 chart: dict | None, path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "报告"
    header_fill = PatternFill("solid", fgColor="1D3557")
    row = 1
    ws.cell(row, 1, title).font = Font(size=16, bold=True)
    row += 1
    ws.cell(row, 1, f"生成时间：{datetime.now().astimezone().isoformat()}")
    row += 2

    for sec in sections:
        ws.cell(row, 1, sec.get("heading", "")).font = Font(
            size=12, bold=True, color="1D3557")
        row += 1
        if sec.get("body"):
            for line in str(sec["body"]).splitlines():
                ws.cell(row, 1, line); row += 1
        tbl = sec.get("table")
        if tbl:
            for ci, h in enumerate(tbl["headers"], 1):
                c = ws.cell(row, ci, h)
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = header_fill
                c.alignment = Alignment(horizontal="center")
            row += 1
            for r in tbl["rows"]:
                for ci, v in enumerate(r, 1):
                    ws.cell(row, ci, v)
                row += 1
        row += 1

    if chart:
        ws2 = wb.create_sheet("趋势数据")
        ws2.cell(1, 1, chart.get("title", ""))
        headers = ["时间"] + [s.get("name", "") for s in chart.get("series", [])]
        for ci, h in enumerate(headers, 1):
            ws2.cell(2, ci, h).font = Font(bold=True)
        for ri, x in enumerate(chart.get("x", []), 3):
            ws2.cell(ri, 1, x)
            for si, s in enumerate(chart.get("series", []), 2):
                ws2.cell(ri, si, s["data"][ri - 3]
                         if ri - 3 < len(s["data"]) else None)

    for w in wb.worksheets:
        for col in w.columns:
            w.column_dimensions[col[0].column_letter].width = 18
    wb.save(path)


# ───────────────────────── CSV ─────────────────────────
def _build_csv(title: str, sections: list[dict],
               chart: dict | None, path: Path) -> None:
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow([title])
        w.writerow([f"生成时间：{datetime.now().astimezone().isoformat()}"])
        w.writerow([])
        for sec in sections:
            w.writerow([f"## {sec.get('heading', '')}"])
            if sec.get("body"):
                for line in str(sec["body"]).splitlines():
                    w.writerow([line])
            tbl = sec.get("table")
            if tbl:
                w.writerow(tbl["headers"])
                w.writerows(tbl["rows"])
            w.writerow([])
        if chart:
            w.writerow([f"## {chart.get('title', '')}"])
            w.writerow(["时间"] + [s.get("name", "")
                                  for s in chart.get("series", [])])
            for i, x in enumerate(chart.get("x", [])):
                w.writerow([x] + [s["data"][i] if i < len(s["data"]) else ""
                                  for s in chart.get("series", [])])


_BUILDERS = {"pdf": _build_pdf, "excel": _build_excel, "csv": _build_csv}
_EXT = {"pdf": "pdf", "excel": "xlsx", "csv": "csv"}


async def export_report(fmt: str, title: str, sections: list[dict],
                        chart: dict | None = None) -> dict:
    if fmt not in _BUILDERS:
        raise ValueError(f"不支持的导出格式: {fmt}")
    trace_id = trace_id_var.get() or uuid.uuid4().hex[:8]
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = REPORT_DIR / f"{trace_id}_{ts}.{_EXT[fmt]}"
    await asyncio.to_thread(_BUILDERS[fmt], title, sections, chart, path)
    return _register(path)